from openpyxl import load_workbook, Workbook

class ExcelHandler:
    def __init__(self, config):
        """ Initialize input and output Excel paths from config. """
        self.input_file = config["input_file"]
        self.output_file = config["output_file"]
        self.output_sheet_name = config["output_sheet_name"]
        self.columns = config["columns"]

    def read_input_data(self):
        """ Read input Excel file and return data as a list of rows. """
        workbook = load_workbook(self.input_file)
        sheet = workbook.active
        return list(sheet.iter_rows(min_row=2, max_col=3, values_only=True))  # Skip header row

    def write_output_data(self, data):
        """ Write scraped data to an output Excel file. """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.output_sheet_name
        sheet.append(self.columns)  # Add headers

        for row in data:
            sheet.append(row)

        workbook.save(self.output_file)
